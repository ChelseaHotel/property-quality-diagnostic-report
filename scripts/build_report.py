#!/usr/bin/env python3
"""Preflight and build an offline property quality diagnostic report."""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import re
import secrets
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl


FIELD_ALIASES = {
    "id": ["整改单号", "工单号", "整改编号", "记录编号", "单号"],
    "project": ["项目", "项目名称", "小区", "小区名称", "服务项目"],
    "status": ["整改单状态", "整改状态", "处理状态", "工单状态", "状态"],
    "task": ["品质检查任务名称", "品质检查任务", "检查任务", "检查批次", "任务名称"],
    "problem": ["问题描述", "问题内容", "不合格项", "现场问题", "问题"],
    "board": ["所属板块", "业务板块", "责任板块", "专业板块", "专业条线"],
    "dimension": ["检查维度", "检查项", "检查类别", "检查项目", "问题类别"],
    "score": ["扣罚分值", "扣分", "处罚分值", "扣罚分", "分值"],
    "measure": ["整改措施", "处理措施", "整改方案"],
    "implementation": ["实施描述", "整改说明", "处理说明", "完成情况"],
    "submitter": ["提交人", "检查人", "创建人", "发起人"],
    "submitted_at": ["提交时间", "创建时间", "检查时间", "发起时间", "日期"],
}
DONE_STATUSES = ["已整改", "已完成", "已关闭", "完成", "关闭"]
PENDING_STATUSES = ["待整改", "整改中", "未完成", "处理中", "待处理", "未整改"]
NO_ISSUE_VALUES = ["", "无", "无。", "无问题", "正常", "符合", "未发现问题", "暂无"]
REQUIRED_FIELDS = ("status", "problem", "board")


def norm(value) -> str:
    return re.sub(r"[\s　_\-—:：()（）【】\[\]/\\]+", "", "" if value is None else str(value)).lower()


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    return str(value).strip()


def load_config(path: str | None) -> dict:
    raw = {}
    if path:
        raw = json.loads(Path(path).read_text(encoding="utf-8-sig"))
    aliases = {key: list(values) for key, values in FIELD_ALIASES.items()}
    for key, values in raw.get("field_aliases", {}).items():
        if key in aliases and isinstance(values, list):
            aliases[key] = list(dict.fromkeys([*values, *aliases[key]]))
    return {
        "raw": raw,
        "aliases": aliases,
        "done_statuses": list(dict.fromkeys(raw.get("done_statuses", []) + DONE_STATUSES)),
        "pending_statuses": list(dict.fromkeys(raw.get("pending_statuses", []) + PENDING_STATUSES)),
        "no_issue_values": list(dict.fromkeys(raw.get("no_issue_values", []) + NO_ISSUE_VALUES)),
    }


def alias_score(header, aliases: list[str]) -> int:
    token = norm(header)
    if not token:
        return -1
    best = -1
    for alias in aliases:
        candidate = norm(alias)
        if token == candidate:
            best = max(best, 100)
        elif candidate and candidate in token:
            best = max(best, 70 + min(20, len(candidate) * 2))
        elif token in candidate and len(token) >= 2:
            best = max(best, 55)
    return best


def detect_header(row: tuple, aliases: dict) -> tuple[dict, int, list[str]]:
    mapping = {}
    confidence_warnings = []
    score = 0
    for field, field_aliases in aliases.items():
        candidates = sorted(
            ((alias_score(value, field_aliases), index, clean(value)) for index, value in enumerate(row)),
            reverse=True,
        )
        candidates = [item for item in candidates if item[0] >= 55]
        mapping[field] = candidates[0][1] if candidates else None
        if candidates:
            score += 3 if field in REQUIRED_FIELDS else 1
            if len(candidates) > 1 and candidates[0][0] - candidates[1][0] < 8:
                confidence_warnings.append(f"字段 {field} 存在相近候选：{candidates[0][2]} / {candidates[1][2]}")
    return mapping, score, confidence_warnings


def choose_sheet_and_header(workbook, config: dict, sheet_name: str | None, header_row: int | None):
    sheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
    candidates = []
    for sheet in sheets:
        rows = [header_row] if header_row else range(1, min(sheet.max_row, 20) + 1)
        for row_number in rows:
            values = tuple(cell.value for cell in sheet[row_number])
            mapping, score, warnings = detect_header(values, config["aliases"])
            candidates.append((score, sheet.title, row_number, mapping, warnings, [clean(v) for v in values]))
    if not candidates:
        raise ValueError("未找到可分析的工作表")
    candidates.sort(key=lambda item: (-item[0], item[1], item[2]))
    return candidates[0], candidates[:5]


def parse_score(value) -> tuple[float, bool]:
    if value in (None, ""):
        return 0.0, True
    text = clean(value).replace("分", "").strip()
    try:
        return float(text), True
    except ValueError:
        return 0.0, False


def status_kind(status: str, config: dict) -> str:
    token = norm(status)
    if token in {norm(value) for value in config["done_statuses"]}:
        return "done"
    if token in {norm(value) for value in config["pending_statuses"]}:
        return "pending"
    return "unknown"


def load_rows(input_path: str, args, config: dict) -> tuple[list[dict], dict]:
    workbook = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    try:
        best, ranked = choose_sheet_and_header(workbook, config, args.sheet, args.header_row)
        score, sheet_name, header_row, mapping, header_warnings, header_values = best
        sheet = workbook[sheet_name]
        missing = [field for field in REQUIRED_FIELDS if mapping.get(field) is None]
        rows = []
        invalid_score_rows = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not any(value is not None and clean(value) for value in values):
                continue
            def get(field):
                index = mapping.get(field)
                return values[index] if index is not None and index < len(values) else None
            score_value, score_valid = parse_score(get("score"))
            if not score_valid:
                invalid_score_rows.append(row_number)
            rows.append({
                "id": clean(get("id")),
                "project": clean(get("project")),
                "status": clean(get("status")),
                "status_kind": status_kind(clean(get("status")), config),
                "task": clean(get("task")),
                "problem": clean(get("problem")),
                "board": clean(get("board")),
                "dimension": clean(get("dimension")),
                "score": score_value,
                "measure": clean(get("measure")),
                "implementation": clean(get("implementation")),
                "submitter": clean(get("submitter")),
                "submitted_at": clean(get("submitted_at")),
                "source_row": row_number,
            })
    finally:
        workbook.close()
    metadata = {
        "sheet": sheet_name,
        "header_row": header_row,
        "header_score": score,
        "header": header_values,
        "mapping": {key: (header_values[index] if index is not None and index < len(header_values) else None) for key, index in mapping.items()},
        "missing_required_fields": missing,
        "header_warnings": header_warnings,
        "invalid_score_rows": invalid_score_rows,
        "ranked_headers": [{"sheet": item[1], "row": item[2], "score": item[0]} for item in ranked],
    }
    return rows, metadata


def is_issue(record: dict, policy: str, config: dict) -> bool:
    problem_valid = norm(record["problem"]) not in {norm(value) for value in config["no_issue_values"]}
    if policy == "problem-only":
        return problem_valid
    if policy == "score-only":
        return record["score"] > 0
    if policy == "status-based":
        return record["status_kind"] in {"done", "pending"}
    return problem_valid or record["score"] > 0


def filename_project_candidate(path: str) -> str:
    stem = Path(path).stem
    generic_stem = norm(stem)
    if re.fullmatch(r"(?:测试|test|数据|data|导出|export|报告|report|样例|示例|sample)\d*", generic_stem, re.I):
        return ""
    if not re.search(r"[\u4e00-\u9fff]", stem):
        return ""
    stem = re.split(r"\d{4}|\d{1,2}月", stem, maxsplit=1)[0]
    stem = re.sub(r"(品质|品检|检查|整改|记录|明细|结果|报告)+$", "", stem)
    return stem.strip(" _-—")


def build_preflight(input_path: str, rows: list[dict], metadata: dict, args, config: dict) -> dict:
    projects = Counter(record["project"] for record in rows if record["project"])
    statuses = Counter(record["status"] or "空白" for record in rows)
    unknown_statuses = sorted({record["status"] or "空白" for record in rows if record["status_kind"] == "unknown"})
    ids = [record["id"] for record in rows if record["id"]]
    duplicate_ids = sorted(value for value, count in Counter(ids).items() if count > 1)
    candidate = filename_project_candidate(input_path)
    reasons = []
    warnings = list(metadata["header_warnings"])
    if metadata["missing_required_fields"]:
        reasons.append("缺少必需字段：" + ", ".join(metadata["missing_required_fields"]))
    if len(projects) > 1 and not args.project:
        reasons.append("检测到多个项目，必须使用 --project 选择")
    if duplicate_ids:
        reasons.append(f"检测到 {len(duplicate_ids)} 个重复工单号")
    if unknown_statuses:
        reasons.append("存在未知整改状态：" + "、".join(unknown_statuses))
    if metadata["invalid_score_rows"]:
        reasons.append(f"存在 {len(metadata['invalid_score_rows'])} 行无法解析的扣分")
    if candidate and len(projects) == 1:
        project = next(iter(projects))
        if norm(candidate) not in norm(project) and norm(project) not in norm(candidate):
            reasons.append(f"文件名候选项目“{candidate}”与表内项目“{project}”不一致")
    if args.privacy == "public" and not args.synthetic:
        reasons.append("public 模式只允许合成数据，必须显式传入 --synthetic")
    issue_count = sum(is_issue(record, args.issue_policy, config) for record in rows)
    if not rows:
        reasons.append("没有可用记录")
    if rows and not issue_count:
        reasons.append("按当前问题口径没有实际问题")
    sensitive_fields = [field for field in ("id", "project", "submitter", "problem") if any(record[field] for record in rows)]
    return {
        "input_file": str(Path(input_path).resolve()),
        "sheet": metadata["sheet"],
        "header_row": metadata["header_row"],
        "field_mapping": metadata["mapping"],
        "projects": dict(projects),
        "statuses": dict(statuses),
        "records": len(rows),
        "actual_issues": issue_count,
        "duplicate_ids": duplicate_ids[:20],
        "invalid_score_rows": metadata["invalid_score_rows"][:20],
        "sensitive_fields": sensitive_fields,
        "privacy": args.privacy,
        "issue_policy": args.issue_policy,
        "warnings": warnings,
        "confirmation_required": bool(reasons),
        "confirmation_reasons": reasons,
    }


def redact_text(text: str) -> str:
    value = clean(text)
    value = re.sub(r"(?<!\d)1[3-9]\d{9}(?!\d)", "[手机号]", value)
    value = re.sub(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", "[邮箱]", value)
    value = re.sub(r"(?<!\d)\d{17}[\dXx](?!\d)", "[身份证号]", value)
    value = re.sub(r"(?i)\b[A-Z]?\d{1,4}\s*车位\b|\b车位\s*[A-Z]?\d{1,4}\b", "[车位]", value)
    value = re.sub(r"\d{1,3}\s*(?:号)?\s*楼", "[楼栋]", value)
    value = re.sub(r"\d{1,3}\s*栋", "[楼栋]", value)
    value = re.sub(r"\d{1,3}\s*单元", "[单元]", value)
    value = re.sub(r"\b\d{3,5}\s*(?:室|房)\b", "[房号]", value)
    return value


def anonymize(records: list[dict], privacy: str) -> tuple[list[dict], str | None]:
    if privacy == "local":
        return records, None
    salt = os.environ.get("QUALITY_REPORT_SALT")
    warning = None
    if not salt:
        salt = secrets.token_hex(16)
        warning = "未设置 QUALITY_REPORT_SALT；本次匿名编号无法跨报告稳定关联。"
    people = {}
    output = []
    for record in records:
        item = dict(record)
        if item["id"]:
            digest = hashlib.sha256((salt + item["id"]).encode("utf-8")).hexdigest()[:8].upper()
            item["id"] = f"ISSUE-{digest}"
        if item["submitter"]:
            people.setdefault(item["submitter"], f"人员-{len(people)+1:02d}")
            item["submitter"] = people[item["submitter"]]
        for field in ("problem", "measure", "implementation"):
            item[field] = redact_text(item[field])
        if privacy == "public":
            item["project"] = "示例项目"
        output.append(item)
    return output, warning


def safe_filename(value: str) -> str:
    value = re.sub(r"[<>:\"/\\|?*\x00-\x1f]", "", value).strip(" .")
    return value or "品质诊断报告"


def build_context(rows: list[dict], input_path: str, args, config: dict, preflight: dict) -> tuple[dict, list[dict]]:
    selected = rows
    if args.project:
        selected = [record for record in selected if record["project"] == args.project]
    if not selected:
        raise ValueError("项目筛选后没有数据")
    issues = [record for record in selected if is_issue(record, args.issue_policy, config)]
    if not issues:
        raise ValueError("按当前问题口径没有实际问题")
    done = [record for record in issues if record["status_kind"] == "done"]
    pending = [record for record in issues if record["status_kind"] == "pending"]
    project_values = [record["project"] for record in selected if record["project"]]
    project = args.project_name or (Counter(project_values).most_common(1)[0][0] if project_values else "物业项目")
    if args.privacy == "public":
        project = "示例项目"
    board_stats = []
    for board, records_in_board in sorted(defaultdict(list, {board: [r for r in issues if r["board"] == board] for board in {r["board"] for r in issues}}).items()):
        board_done = sum(record["status_kind"] == "done" for record in records_in_board)
        board_pending = sum(record["status_kind"] == "pending" for record in records_in_board)
        board_stats.append({
            "board": board or "未分类",
            "issues": len(records_in_board),
            "done": board_done,
            "pending": board_pending,
            "rate": round(board_done / len(records_in_board) * 100, 1),
            "score": round(sum(record["score"] for record in records_in_board), 2),
            "pending_score": round(sum(record["score"] for record in records_in_board if record["status_kind"] == "pending"), 2),
        })
    batches = []
    batch_key = "task" if any(record["task"] for record in selected) else "submitted_at"
    groups = defaultdict(list)
    for record in selected:
        key = record[batch_key] or "未标注批次"
        if batch_key == "submitted_at":
            key = key[:10] or "未标注日期"
        groups[key].append(record)
    for label, batch_rows in groups.items():
        batch_issues = [record for record in batch_rows if is_issue(record, args.issue_policy, config)]
        batch_done = sum(record["status_kind"] == "done" for record in batch_issues)
        batches.append({
            "label": label,
            "checks": len(batch_rows),
            "issues": len(batch_issues),
            "done": batch_done,
            "pending": sum(record["status_kind"] == "pending" for record in batch_issues),
            "rate": round(batch_done / len(batch_issues) * 100, 1) if batch_issues else 0,
            "score": round(sum(record["score"] for record in batch_issues), 2),
        })
    lowest = min(board_stats, key=lambda item: (item["rate"], -item["issues"]))
    pending_heavy = max(board_stats, key=lambda item: (item["pending"], item["pending_score"]))
    highest_pending = max(pending, key=lambda record: record["score"], default=None)
    findings = [
        {"value": f"{lowest['rate']:.1f}%", "title": f"{lowest['board']}整改率最低", "text": f"{lowest['issues']} 条实际问题中仍有 {lowest['pending']} 条待整改。", "risk": lowest["rate"] < 70},
        {"value": str(pending_heavy["pending"]), "title": f"{pending_heavy['board']}待整改最多", "text": f"未关闭问题关联扣分 {pending_heavy['pending_score']:g} 分。", "risk": pending_heavy["pending"] > 0},
    ]
    if highest_pending:
        findings.append({"value": f"{highest_pending['score']:g}分", "title": "单条未整改问题扣分最高", "text": redact_text(highest_pending["problem"]) if args.privacy != "local" else highest_pending["problem"], "risk": highest_pending["score"] >= 3})
    dates = sorted(record["submitted_at"][:10] for record in selected if record["submitted_at"])
    period = f"{dates[0]}—{dates[-1]}" if dates else "未提供日期"
    metrics = {
        "records": len(selected),
        "issues": len(issues),
        "done": len(done),
        "pending": len(pending),
        "rate": round(len(done) / len(issues) * 100, 1),
        "all_rate": round(sum(record["status_kind"] == "done" for record in selected) / len(selected) * 100, 1),
        "score": round(sum(record["score"] for record in issues), 2),
        "pending_score": round(sum(record["score"] for record in pending), 2),
    }
    redacted_rows, privacy_warning = anonymize(selected, args.privacy)
    title = args.title or f"{project}品质诊断报告"
    source_label = Path(input_path).name if args.privacy == "local" else ("内部数据源" if args.privacy == "safe" else "synthetic-input.xlsx")
    context = {
        "project": project,
        "title": title,
        "period": period,
        "privacy": args.privacy,
        "privacy_warning": privacy_warning,
        "source_label": source_label,
        "issue_policy": args.issue_policy,
        "metrics": metrics,
        "boards": board_stats,
        "batches": batches,
        "findings": findings,
        "preflight_warnings": preflight["confirmation_reasons"] if args.confirm_preflight else preflight["warnings"],
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    return context, redacted_rows


def render_report(context: dict, records: list[dict], output_path: Path) -> None:
    template_path = Path(__file__).resolve().parents[1] / "assets" / "report-template.html"
    template = template_path.read_text(encoding="utf-8")
    payload = json.dumps(records, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    context_json = json.dumps(context, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    report = template.replace("__REPORT_TITLE__", html.escape(context["title"]))
    report = report.replace("__DATA__", payload).replace("__CONTEXT__", context_json)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(report, encoding="utf-8")


def make_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="生成物业品质诊断 HTML 报告")
    parser.add_argument("input", help="输入 Excel 文件")
    parser.add_argument("--out", default="outputs", help="输出目录")
    parser.add_argument("--preflight-only", action="store_true")
    parser.add_argument("--confirm-preflight", action="store_true")
    parser.add_argument("--sheet")
    parser.add_argument("--header-row", type=int)
    parser.add_argument("--project")
    parser.add_argument("--project-name")
    parser.add_argument("--title")
    parser.add_argument("--config")
    parser.add_argument("--privacy", choices=("local", "safe", "public"), default="local")
    parser.add_argument("--synthetic", action="store_true", help="确认输入完全由合成数据构成")
    parser.add_argument("--issue-policy", choices=("problem-or-score", "problem-only", "score-only", "status-based"), default="problem-or-score")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = make_parser().parse_args(argv)
    config = load_config(args.config)
    try:
        rows, metadata = load_rows(args.input, args, config)
        preflight = build_preflight(args.input, rows, metadata, args, config)
        print(json.dumps(preflight, ensure_ascii=False, indent=2))
        if args.preflight_only:
            return 0
        if preflight["confirmation_required"] and not args.confirm_preflight:
            print("预检存在需确认风险；修正参数后重试，或人工核实后使用 --confirm-preflight。", file=sys.stderr)
            return 2
        context, output_rows = build_context(rows, args.input, args, config, preflight)
        output = Path(args.out) / f"{safe_filename(context['project'])}品质诊断报告.html"
        render_report(context, output_rows, output)
        print(json.dumps({"output": str(output.resolve()), "metrics": context["metrics"], "privacy_warning": context["privacy_warning"]}, ensure_ascii=False, indent=2))
        return 0
    except (ValueError, FileNotFoundError, KeyError, json.JSONDecodeError) as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
